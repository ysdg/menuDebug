# !/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from dataTransfer import *
from datetime import datetime
from re import findall as reFindall

class excelWrite(openpyxl.Workbook):
	""" 
	excel writing operation based on openpyxl.Workbook:
		excelColName: default excel column name;
		excelName: name to store, including machine code and datatime;
		align: default alignment, right for horizontal, center for vertical;
		sheetIndex: default sheet index, will add for adding new sheet;
		excelRelativePath: relative path to store excel.
	 """
	def __init__(self, machineName, menuNameDict):
		super(excelWrite, self).__init__()
		print("open excel work book successfully.")
		self.machineName = machineName
		self.menuNameDict = menuNameDict
		self.excelColName = ['步骤', '功能', '步骤时间', '倒计时', '备注']
		self.excelName = self.machineName+"菜单流程"+datetime.now().strftime('%Y%m%d_%H%M%S')+".xlsx"
		self.align = openpyxl.styles.Alignment(horizontal='right',vertical='center',wrap_text=True)
		borderSide = openpyxl.styles.Side(color="FF000000", style="thin")
		self.border = openpyxl.styles.Border(left=borderSide, right=borderSide, top=borderSide, bottom=borderSide)
		self.sheetIndex = 0
		self.excelRelativePath = "./"
	def saveExcel(self):
		""" 
		save excel with excelName.
		 """
		self.save(self.excelName)
		print("save excel: "+self.excelName+", successfully.")
	def createNewSheet(self, nameKey):
		""" 
		create new sheet and reset curSheetRow to 1.
		 """
		self.remainTimeProcess()
		self.curSheet = self.create_sheet(self.menuNameDict[nameKey])
		self.curSheetRow = 1
		print("Create sheet: "+self.menuNameDict[nameKey]+", successfully.")
		if self.sheetnames[0]=="Sheet": self.remove(self["Sheet"])
	def rowAlignSet(self):
		""" 
		set row alignment.
		 """
		self.curSheet['A'+str(self.curSheetRow)].alignment 	= self.align
		self.curSheet['B'+str(self.curSheetRow)].alignment 	= self.align
		self.curSheet['C'+str(self.curSheetRow)].alignment 	= self.align
		self.curSheet['D'+str(self.curSheetRow)].alignment 	= self.align
		self.curSheet['E'+str(self.curSheetRow)].alignment 	= self.align
		self.curSheet['A'+str(self.curSheetRow)].border 	= self.border
		self.curSheet['B'+str(self.curSheetRow)].border 	= self.border
		self.curSheet['C'+str(self.curSheetRow)].border 	= self.border
		self.curSheet['D'+str(self.curSheetRow)].border 	= self.border
		self.curSheet['E'+str(self.curSheetRow)].border 	= self.border
		self.curSheet.column_dimensions['A'].width = 10
		self.curSheet.column_dimensions['B'].width = 30
		self.curSheet.column_dimensions['C'].width = 15
		self.curSheet.column_dimensions['D'].width = 15
		self.curSheet.column_dimensions['E'].width = 30
	def dataRecode(self, dat):
		""" 
		recode from line dat to writing row dat;
		no remain time.
		 """
		step = self.curSheetRow-2
		if 	dat[-1].find('Temp') != -1 or \
			dat[-1].find('TEMP') != -1 or \
			dat[-1].find('temp') != -1: 
				dat[-1] = reFindall(r"\d+\.?\d*", dat[-1])[0]
		if  dat[1].find("Temp")!=-1 or \
			dat[1].find("TEMP")!=-1 or \
			dat[1].find("temp")!=-1:
			if dat[0] in list(workHeadDictH)[5:10]:
				if dat[-2] in list(heatRankDict):
					ctrlMess = heatRankDict[dat[-2]]+'档加热至'+dat[-1]+'度'+'\n'+'('+workHeadDictH[dat[0]]+')'
			elif dat[0]==list(workHeadDictH)[2]:
				ctrlMess = heatRankDict[dat[-2]]+'档'+workHeadDictH[dat[0]]+'至'+dat[-1]+'度'
			elif dat[0]==list(workHeadDictH)[1]:
				ctrlMess = heatRankDict[dat[-1]]+'档'+workHeadDictH[dat[0]]
			elif dat[0]==list(workHeadDictH)[0]:
				ctrlMess = workHeadDictH[dat[0]]
			elif dat[0]==list(workHeadDictH)[4]:
				ctrlMess = workHeadDictH[dat[0]]+'上'+dat[-2]+'到'+dat[-1]+'步'
			else: ctrlMess = dat[0]
		else:
			if dat[0] in list(workHeadDictH)[5:10]:
				if dat[-2] in list(heatRankDict):
					ctrlMess = heatRankDict[dat[-2]]+'档加热'+'\n'+'('+workHeadDictH[dat[0]]+')'
			elif dat[0]==list(workHeadDictH)[2]:
				ctrlMess = heatRankDict[dat[-2]]+'档'+workHeadDictH[dat[0]]
			elif dat[0]==list(workHeadDictH)[1]:
				ctrlMess = heatRankDict[dat[-1]]+'档'+workHeadDictH[dat[0]]
			elif dat[0]==list(workHeadDictH)[0]:
				ctrlMess = workHeadDictH[dat[0]]
			elif dat[0]==list(workHeadDictH)[4]:
				ctrlMess = workHeadDictH[dat[0]]+'上'+dat[-2]+'到'+dat[-1]+'步'
			else: ctrlMess = workHeadDictH[dat[0]]

		if int(dat[3]) >= 60: 
			stepTime = '1:'+str(int(dat[3])-60).zfill(2)+':'+str(int(dat[4])).zfill(2)
		else: 
			stepTime = '0:'+str(int(dat[3])).zfill(2)+':'+str(int(dat[4])).zfill(2)
		remainTime = "0:00:00"
		endCondition = '以'+workHeadDictL[dat[1]]
		if dat[-1]=='DISP_UPDATE': endCondition = endCondition+"(更新时间)"
		return [step, ctrlMess, stepTime, remainTime, endCondition]
	def remainTimeProcess(self):
		""" 
		count and writing remain time of process.
		 """
		timeFormat = "%H:%M:%S"
		baseTime = datetime.strptime("0:00:00", timeFormat)
		try: 
			rowMax = self.curSheetRow-2
			row = self.curSheetRow-2
		except: return 
		while row > 2:
			if rowMax==row:
				remainTimeStr = self.curSheet.cell(row = row, column=3).value
				remainTime = datetime.strptime(remainTimeStr, timeFormat)
			else:
				remainTimeStr = self.curSheet.cell(row = row, column=3).value
				remainTime = datetime.strptime(remainTimeStr, timeFormat)
				remainTimeStr = self.curSheet.cell(row = row+1, column=4).value
				remainTime = remainTime + (datetime.strptime(remainTimeStr, timeFormat)-baseTime)
			self.curSheet.cell(row=row, column=4, value=remainTime.strftime(timeFormat))
			row = row-1
	def writeRowData(self, dat):
		""" 
		write row data to excel after recoding.
		 """
		if  dat == ['0', '0', '0', '0', '0'] or dat==[]: 
			self.curSheetRow = self.curSheetRow+1
			return 
		if self.curSheetRow==1:
			self.curSheet.row_dimensions[1].height = 30
			self.curSheet['A1'] = '菜单'
			self.curSheet['B1'] = self.curSheet.title
			self.curSheet['C1'] = "适用机型"
			self.curSheet['D1'] = self.machineName
			self.curSheet.cell(self.curSheetRow, column=1).font = openpyxl.styles.Font('微软雅黑', 12, True)
			self.curSheet.cell(self.curSheetRow, column=2).font = openpyxl.styles.Font('微软雅黑', 12, True)
			self.curSheet.cell(self.curSheetRow, column=3).font = openpyxl.styles.Font('微软雅黑', 12, True)
			self.curSheet.cell(self.curSheetRow, column=4).font = openpyxl.styles.Font('微软雅黑', 12, True)
			self.rowAlignSet()
			self.curSheetRow = 2
			self.curSheet.append(self.excelColName)
			self.rowAlignSet()
			self.curSheetRow = self.curSheetRow+1
		self.curSheet.append(self.dataRecode(dat))
		self.curSheet.cell(self.curSheetRow, column=3).number_format = openpyxl.styles.numbers.FORMAT_DATE_TIME6
		self.curSheet.cell(self.curSheetRow, column=4).number_format = openpyxl.styles.numbers.FORMAT_DATE_TIME6
		self.rowAlignSet()
		self.curSheetRow = self.curSheetRow+1


if __name__=="__main__":
	pass

